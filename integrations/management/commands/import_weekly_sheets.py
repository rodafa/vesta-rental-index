import logging
import os
from datetime import datetime, timedelta

import openpyxl

from django.core.management.base import BaseCommand
from django.db.models import Q

from dashboard.models import PropertyWeeklyNote
from market.models import WeeklyLeasingSummary
from properties.models import Unit

logger = logging.getLogger(__name__)


def _safe_ratio(numerator, denominator):
    if not denominator:
        return None
    return round(numerator / denominator * 100, 2)


def _week_monday(d):
    """Given a date, return the Monday of that week."""
    return d - timedelta(days=d.weekday())


def _match_unit(property_name, unit_cache):
    """Match a property name from Google Sheet to a Unit in the database.

    Returns (Unit, match_description) or (None, reason).
    """
    if property_name in unit_cache:
        return unit_cache[property_name]

    name = property_name.strip().strip('"').strip("'").strip()
    name_lower = name.lower()

    def _pick_best(matches, label):
        """When multiple units match an address, disambiguate by unit name keywords."""
        import re as _re

        if matches.count() == 1:
            unit = matches.first()
            return unit, f"{label}: {unit.property.address_line_1}"

        if matches.count() > 1:
            units_list = list(matches)

            # Try matching "unit X" or trailing single letter/number against unit names
            # e.g. "588 Ray Hill Road D" → trailing "D" matches unit name "D"
            trailing_m = _re.search(r'\s+([a-z])$', name_lower)
            if trailing_m:
                token = trailing_m.group(1)
                candidates = [u for u in units_list if (u.name or "").lower().strip() == token]
                if len(candidates) == 1:
                    return candidates[0], f"{label} (trailing letter '{token}'): {candidates[0].property.address_line_1}"

            m = _re.search(r'unit\s+([a-z0-9]+)', name_lower)
            if m:
                token = m.group(1)
                candidates = [u for u in units_list if token in (u.name or "").lower()]
                if len(candidates) == 1:
                    return candidates[0], f"{label} (unit '{token}' name match): {candidates[0].property.address_line_1}"

                # For numeric tokens, try ordering by rentengine_id
                if token.isdigit():
                    unit_num = int(token)
                    ordered = sorted(units_list, key=lambda u: u.rentengine_id or 0)
                    idx = unit_num - 1
                    if 0 <= idx < len(ordered):
                        unit = ordered[idx]
                        return unit, f"{label} (unit {unit_num} of {len(ordered)} by rentengine_id): {unit.property.address_line_1}"

            # Try matching any standalone number from the name against unit names.
            # e.g. "701 Macedonia Road 707" → "707" matches unit with name "707".
            # Skip the first number if it matches the street address (not a unit identifier).
            addr_prefix = (units_list[0].property.address_line_1 or "").split()[0] if units_list else ""
            for token in _re.findall(r'\b(\d{2,})\b', name_lower):
                if token == addr_prefix:
                    continue  # skip street number
                candidates = [u for u in units_list if token in (u.name or "").lower()]
                if len(candidates) == 1:
                    return candidates[0], f"{label} (number '{token}' in unit name): {candidates[0].property.address_line_1}"

            # Try to match positional keywords (top/bottom/upper/lower)
            keyword_map = {
                "bottom": ["bottom", "lower", "downstairs"],
                "top": ["top", "upper", "upstairs"],
            }
            for db_kw, sheet_kws in keyword_map.items():
                if any(kw in name_lower for kw in sheet_kws):
                    candidates = [u for u in units_list if db_kw in (u.name or "").lower()]
                    if len(candidates) == 1:
                        return candidates[0], f"{label} (kw '{db_kw}'): {candidates[0].property.address_line_1}"

            # Default to first match
            unit = units_list[0]
            return unit, f"{label} first of {len(units_list)}: {unit.property.address_line_1}"

        return None, None

    # Build search terms: full name, 3-word prefix, reversed-number variants, 2-word prefix.
    # Order matters: more specific before looser.
    parts = name.split()
    search_terms = [name]
    if len(parts) >= 3:
        search_terms.append(" ".join(parts[:3]))

    # If last significant token is a number, try "number + first word(s)" BEFORE the
    # 2-word fallback so e.g. "Bartlett St 178" tries "178 Bartlett St" before "Bartlett St".
    last = parts[-1].rstrip("_.,")
    if last.isdigit():
        num = last
        rest = parts[:-1]
        if rest:
            if len(rest) >= 2:
                search_terms.append(num + " " + " ".join(rest[:2]))
            search_terms.append(num + " " + rest[0])

    if len(parts) >= 2:
        search_terms.append(" ".join(parts[:2]))

    # Handle "Street Name 123 ..." — number embedded anywhere in name.
    # Extract first 3+ digit number from name, try "number + first word".
    import re as _re2
    m_mid = _re2.search(r'\b(\d{2,})\b', name)
    if m_mid:
        num_mid = m_mid.group(1)
        first_word = parts[0]
        if num_mid != first_word:  # don't duplicate if number is already first
            search_terms.append(num_mid + " " + first_word)
            if len(parts) >= 2:
                search_terms.append(num_mid + " " + " ".join(parts[:2]))

    for term in search_terms:
        matches = Unit.objects.filter(
            Q(property__address_line_1__icontains=term) | Q(address_line_1__icontains=term),
            is_active=True,
        ).select_related("property")

        if matches.exists():
            unit, desc = _pick_best(matches, f"match on '{term}'")
            if unit:
                unit_cache[property_name] = (unit, desc)
                return unit_cache[property_name]

    # Try without is_active filter (for closed-out listings)
    for term in search_terms:
        matches = Unit.objects.filter(
            Q(property__address_line_1__icontains=term) | Q(address_line_1__icontains=term),
        ).select_related("property")

        if matches.exists():
            unit, desc = _pick_best(matches, f"inactive match on '{term}'")
            if unit:
                unit_cache[property_name] = (unit, desc)
                return unit_cache[property_name]

    unit_cache[property_name] = (None, f"no match found for '{name}'")
    return unit_cache[property_name]


def _parse_cell_date(value):
    """Parse a cell value into a date. Handles datetime objects and strings."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year"):  # date object
        return value
    # Try parsing as string
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _safe_int(value):
    """Convert a cell value to int, defaulting to 0."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


class Command(BaseCommand):
    help = "Import weekly leasing data from Google Sheet xlsx exports into WeeklyLeasingSummary and PropertyWeeklyNote"

    def add_arguments(self, parser):
        parser.add_argument(
            "--dir",
            type=str,
            required=True,
            help="Path to directory containing xlsx files (one per property)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to DB",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete existing WeeklyLeasingSummary records before import (fresh start)",
        )

    def handle(self, *args, **options):
        src_dir = options["dir"]
        dry_run = options["dry_run"]
        clear = options["clear"]

        if dry_run:
            self.stdout.write("=== DRY RUN ===\n")

        if not os.path.isdir(src_dir):
            self.stdout.write(self.style.ERROR(f"Directory not found: {src_dir}"))
            return

        xlsx_files = [f for f in os.listdir(src_dir) if f.lower().endswith(".xlsx")]
        if not xlsx_files:
            self.stdout.write(self.style.ERROR(f"No xlsx files found in {src_dir}"))
            return

        self.stdout.write(f"Found {len(xlsx_files)} xlsx files in {src_dir}\n")

        if clear and not dry_run:
            deleted, _ = WeeklyLeasingSummary.objects.all().delete()
            self.stdout.write(self.style.WARNING(f"Cleared {deleted} existing WeeklyLeasingSummary records\n"))

        total_summaries_created = 0
        total_summaries_updated = 0
        total_notes_created = 0
        total_notes_updated = 0
        matched_files = 0
        unmatched = []
        unit_cache = {}

        for xlsx_file in sorted(xlsx_files):
            xlsx_path = os.path.join(src_dir, xlsx_file)
            self.stdout.write(f"\n--- {xlsx_file} ---")

            result = self._process_xlsx(xlsx_path, unit_cache, dry_run)
            if result is None:
                continue

            if result["unit"] is None:
                unmatched.append((xlsx_file, result["property_name"], result["reason"]))
                continue

            matched_files += 1
            total_summaries_created += result["summaries_created"]
            total_summaries_updated += result["summaries_updated"]
            total_notes_created += result["notes_created"]
            total_notes_updated += result["notes_updated"]

        # Summary
        self.stdout.write(f"\n{'='*60}")
        if dry_run:
            self.stdout.write("DRY RUN SUMMARY:")
        else:
            self.stdout.write(self.style.SUCCESS("IMPORT COMPLETE:"))

        self.stdout.write(f"  Files processed: {len(xlsx_files)}")
        self.stdout.write(f"  Properties matched: {matched_files}")
        self.stdout.write(f"  WeeklyLeasingSummary created: {total_summaries_created}")
        self.stdout.write(f"  WeeklyLeasingSummary updated: {total_summaries_updated}")
        self.stdout.write(f"  PropertyWeeklyNote created: {total_notes_created}")
        self.stdout.write(f"  PropertyWeeklyNote updated: {total_notes_updated}")

        if unmatched:
            self.stdout.write(self.style.WARNING(f"\n  Unmatched properties ({len(unmatched)}):"))
            for xlsx_file, prop_name, reason in unmatched:
                self.stdout.write(f"    {xlsx_file}: '{prop_name}' — {reason}")

    def _process_xlsx(self, xlsx_path, unit_cache, dry_run):
        """Process a single xlsx file. Returns dict with results or None on error."""
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"  Error reading file: {e}"))
            return None

        # Use "Data Entry" sheet if available, otherwise first sheet
        if "Data Entry" in wb.sheetnames:
            ws = wb["Data Entry"]
        else:
            ws = wb.active

        # Property name is in cell C1 (column 3, row 1).
        # Some older files have a date range or number there — fall back to filename.
        property_name = None
        c1 = ws.cell(row=1, column=3).value
        if c1 and isinstance(c1, str) and len(c1.strip()) > 4 and not c1.strip()[:1].isdigit():
            property_name = c1.strip()
        if not property_name:
            # Derive from filename: strip .xlsx extension
            property_name = os.path.splitext(os.path.basename(xlsx_path))[0]
        self.stdout.write(f"  Property: {property_name}")

        # Match to unit
        unit, match_info = _match_unit(property_name, unit_cache)
        if unit is None:
            self.stdout.write(self.style.WARNING(f"  {match_info}"))
            wb.close()
            return {"unit": None, "property_name": property_name, "reason": match_info}

        self.stdout.write(f"  Matched: {match_info} (unit_id={unit.id})")

        # Find header row containing "Statement Date" or "Week Ending"
        # New format: B=Statement Date, C=Leads, D=Showings, E=Apps, F=Notes
        # Old format: B=Week Ending In, C=Leads, D=Pre-Qual, E=Sched, F=Showings, G=Apps, H=Notes
        header_row = None
        old_format = False
        for row_num in range(1, 10):
            cell_b = ws.cell(row=row_num, column=2).value
            if cell_b is None:
                continue
            cell_b_str = str(cell_b).lower()
            if "statement" in cell_b_str and "date" in cell_b_str:
                header_row = row_num
                old_format = False
                break
            if "week ending" in cell_b_str:
                header_row = row_num
                old_format = True
                break

        if header_row is None:
            self.stdout.write(self.style.WARNING("  No date header found, skipping"))
            wb.close()
            return {"unit": None, "property_name": property_name, "reason": "no header row"}

        # Column mapping
        # new format: leads=3, showings=4, apps=5, notes=6
        # old format: leads=3, showings=6, apps=7, notes=8
        if old_format:
            col_leads, col_showings, col_apps, col_notes = 3, 6, 7, 8
        else:
            col_leads, col_showings, col_apps, col_notes = 3, 4, 5, 6

        # Parse data rows starting after header
        summaries_created = 0
        summaries_updated = 0
        notes_created = 0
        notes_updated = 0
        rows_parsed = 0

        for row_num in range(header_row + 1, ws.max_row + 1):
            date_val = ws.cell(row=row_num, column=2).value
            week_ending = _parse_cell_date(date_val)
            if week_ending is None:
                continue

            leads = _safe_int(ws.cell(row=row_num, column=col_leads).value)
            showings = _safe_int(ws.cell(row=row_num, column=col_showings).value)
            apps = _safe_int(ws.cell(row=row_num, column=col_apps).value)
            notes_text = ws.cell(row=row_num, column=col_notes).value or ""
            if isinstance(notes_text, str):
                notes_text = notes_text.strip()

            rows_parsed += 1

            if dry_run:
                summaries_created += 1
                if notes_text:
                    notes_created += 1
                continue

            # Write WeeklyLeasingSummary
            _, was_created = WeeklyLeasingSummary.objects.update_or_create(
                week_ending=week_ending,
                unit=unit,
                defaults={
                    "leads_count": leads,
                    "showings_completed_count": showings,
                    "showings_missed_count": 0,
                    "applications_count": apps,
                    "lead_to_show_rate": _safe_ratio(showings, leads),
                    "show_to_app_rate": _safe_ratio(apps, showings),
                    "property_display_name": property_name,
                },
            )
            if was_created:
                summaries_created += 1
            else:
                summaries_updated += 1

            # Write PropertyWeeklyNote if notes present
            if notes_text:
                week_monday = _week_monday(week_ending)
                _, note_created = PropertyWeeklyNote.objects.update_or_create(
                    unit=unit,
                    week_date=week_monday,
                    defaults={
                        "author": "Google Sheet Import",
                        "note_text": notes_text,
                    },
                )
                if note_created:
                    notes_created += 1
                else:
                    notes_updated += 1

        wb.close()
        self.stdout.write(f"  Rows parsed: {rows_parsed}")
        return {
            "unit": unit,
            "property_name": property_name,
            "summaries_created": summaries_created,
            "summaries_updated": summaries_updated,
            "notes_created": notes_created,
            "notes_updated": notes_updated,
        }
